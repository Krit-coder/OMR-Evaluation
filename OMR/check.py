from flask import Flask, render_template, request
import cv2
import numpy as np
import base64
import main
import openpyxl
import os

app = Flask(__name__)


def check_img(img, q, c):
    # num_questions = int(request.form['num_questions'])
    # num_choices = int(request.form['num_choices'])
    final_img = main.check(img, q, c)
    return final_img


def resize_image_to_fit(img, max_width, max_height):
    # Resize the image to fit within the specified maximum width and height
    height, width = img.shape[:2]
    if width > max_width or height > max_height:
        if width > height:
            ratio = max_width / width
        else:
            ratio = max_height / height
        img = cv2.resize(img, (int(width * ratio), int(height * ratio)))
    return img


def update_excel(score):
    # Update Excel sheet with the score
    filename = 'students.xlsx'
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        wb.save(filename)
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    row = sheet.max_row + 1
    sheet.cell(row=row, column=1, value=score)
    wb.save(filename)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/check_sheet', methods=['POST'])
def upload_file():
    uploaded_file = request.files['image']
    num_questions = int(request.form['num_questions'])
    num_choices = int(request.form['num_choices'])
    if uploaded_file.filename != '':
        img_np = np.frombuffer(uploaded_file.read(), np.uint8)
        img = cv2.imdecode(img_np, cv2.IMREAD_COLOR)
        processed_img = check_img(img, num_questions, num_choices)
        # cv2.imshow("Final Image", processed_img)
        # Assume score is calculated here
        ans = main.score  # Example score
        update_excel(ans)
        # Resize the processed image to fit within the window dimensions
        processed_img = resize_image_to_fit(processed_img, 720, 540)
        # Convert processed image to base64 for displaying
        _, processed_img_encoded = cv2.imencode('.png', processed_img)
        processed_img_base64 = base64.b64encode(processed_img_encoded).decode('utf-8')
        return f'''
            <h2>Processed Image</h2>
            <img src="data:image/png;base64,{processed_img_base64}" alt="Processed Image"><br><br>
            <form action="/" method="get">
                <input type="submit" value="Reset">
            </form>
        '''
    return 'No file uploaded'


if __name__ == '__main__':
    app.run(debug=True)
